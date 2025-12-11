from rest_framework import viewsets, permissions, status, filters
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.authtoken.models import Token
from rest_framework.authtoken.views import ObtainAuthToken
from django.contrib.auth.models import User
from .models import Student, MonthlyFee, Attendance, BeltTest, UserRole, StudentEvent
from .serializers import (
    StudentSerializer, 
    MonthlyFeeSerializer, 
    AttendanceSerializer, 
    BeltTestSerializer,
    UserSerializer,
    StudentEventSerializer,
    PasswordChangeSerializer
)

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 1000

class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.filter(is_active=True)
    serializer_class = StudentSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'phone_number', 'instructor_name', 'registration_number']

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()

class MonthlyFeeViewSet(viewsets.ModelViewSet):
    queryset = MonthlyFee.objects.filter(student__is_active=True).order_by('-year', '-month')
    serializer_class = MonthlyFeeSerializer
    permission_classes = [permissions.IsAuthenticated]

class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.filter(student__is_active=True).order_by('-date')
    serializer_class = AttendanceSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    filterset_fields = ['student', 'date']
    
    def get_queryset(self):
        queryset = Attendance.objects.filter(student__is_active=True)
        # We can rely on DjangoFilterBackend for basic filtering, but if we have custom logic for 'date' query param 
        # (which we seem to have handled manually before), we can keep it or delegate to filter backend.
        # The manual check below is redundant if we use filterset_fields = ['date']
        return queryset.order_by('-date')

class BeltTestViewSet(viewsets.ModelViewSet):
    queryset = BeltTest.objects.filter(student__is_active=True).order_by('-test_date')
    serializer_class = BeltTestSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.drawing.image import Image
        from django.http import HttpResponse
        from django.conf import settings
        import os
        from datetime import datetime

        # Belt to number mapping
        def get_belt_number(belt_name):
            belt_map = {
                'white': '10',
                'yellow_stripe': '9',
                'yellow': '8',
                'green_stripe': '7',
                'green': '6',
                'blue_stripe': '5',
                'blue': '4',
                'red_stripe': '3',
                'red': '2',
                'red_black': '1',
            }
            return belt_map.get(belt_name, belt_name)

        queryset = self.filter_queryset(self.get_queryset())

        # Filter inputs
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        status_param = request.query_params.get('status')

        if start_date:
            queryset = queryset.filter(test_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(test_date__lte=end_date)
        if status_param and status_param.lower() != 'all':
            queryset = queryset.filter(result=status_param)

        if not queryset.exists():
            return Response({"error": "No data found."}, status=400)

        # ============================================
        #   RESPONSE
        # ============================================
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="TAI_CERTIFICATE_LIST.xlsx"'

        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "SHEET1"

        # ============================================
        #   STYLES
        # ============================================
        bold = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        info_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        title_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")

        thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
        thick = Border(left=Side(style="thick"), right=Side(style="thick"),
                    top=Side(style="thick"), bottom=Side(style="thick"))

        # ============================================
        #   LOGO – positioned perfectly (NO EXTRA GAP)
        # ============================================
        logo_path = os.path.join(settings.MEDIA_ROOT, "tai_banner.png")
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.width *= 0.90
            img.height *= 0.90
            ws.add_image(img, "F3")   # LOGO MOVED DOWN → FIXED GAP

        # Tight spacing like your sample
        ws.row_dimensions[1].height = 5
        ws.row_dimensions[2].height = 5
        ws.row_dimensions[3].height = 22
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 22

        # ============================================
        #   HEADER BLOCKS – SAME ROW (NO EMPTY SPACE)
        # ============================================
        # Rows 6–8 = Applicant Name, SGB, CLUB
        ws.merge_cells("A7:C7")
        ws.merge_cells("A8:C8")
        ws.merge_cells("A9:C9")

        left_cells = ["A7", "C7", "A8", "C8", "A9", "C9"] 

        left_info = [
            ("A7", "Applicant Name: VENUGOPALAN"),
            ("A8", "SGB : UTAK"),
            ("A9", "CLUB:MASTERS TAEKWON-DO ACADEMY"),
        ]

        for cell_ref, text in left_info:
            c = ws[cell_ref]
            c.value = text
            c.font = bold
            c.fill = info_fill
            c.border = thick
            c.alignment = Alignment(vertical="center")

        for ref in left_cells:
            ws[ref].border = thick

        # RIGHT BLOCK (TAI CERTIFICATE APPLICATION LIST)
        # Add gap between left and right blocks - move to H-K instead of G-I
        ws.merge_cells("G8:I8")

        title_cell = ws["G8"]
        title_cell.value = "TAI CERTIFICATE APPLICATION LIST"
        title_cell.font = bold
        title_cell.fill = title_fill
        title_cell.border = thick
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # These are the left & right edge cells of the merged box
        right_cells = ["G8", "I8"]

        for ref in right_cells:
            ws[ref].border = thick

        # ============================================
        #   COLUMN HEADERS – ROW 10 (LIKE SAMPLE)
        # ============================================
        headers = [
            "Sort",
            "TAI Cert No.",
            "Name of the student",
            "Date of birth",
            "State",
            "Father's Name",
            "Kup Applied for",
            "Date of test",
            "Gender",
            "Age",
            "Instructor Name ( Inst Cert No )",
            "Instructor Affl. No",
            "Last Affl. renewal date",
            "Name of Examiner",
            "Applicant Promoted TAI No.",
            "Last grade Date",
            "TAI Membership Number",
            "FEE",
            "Remarks",
        ]

        header_row = 10
        for col, head in enumerate(headers, 1):
            c = ws.cell(row=header_row, column=col)
            c.value = head
            c.font = bold
            c.fill = header_fill
            c.border = thin
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[header_row].height = 30

        # ============================================
        #   COLUMN WIDTHS – EXACT MATCH TO SAMPLE
        # ============================================
        widths = [
            6,   # Sort
            15,  # TAI Cert No
            28,  # Name
            16,  # DOB
            14,  # State
            22,  # Father
            14,  # Kup
            16,  # Date of test
            10,  # Gender
            8,   # Age
            30,  # Instructor Name
            15,  # Inst Affl No
            22,  # Last Affl
            22,  # Examiner
            22,  # Promoted TAI No
            18,  # Last grade
            24,  # Membership No
            12,  # Fee
            18,  # Remarks
        ]

        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # ============================================
        #   TABLE DATA – STARTING ROW 11
        # ============================================
        for idx, test in enumerate(queryset, 1):
            s = test.student
            row = header_row + idx

            dob = s.date_of_birth.strftime("%-d %b %y") if s.date_of_birth else ""
            test_date = test.test_date.strftime("%d/%m/%Y") if test.test_date else ""

            data = [
                idx,
                s.tai_certification_number or "N/A",
                s.name,
                dob,
                s.state or "",
                s.guardian_name,
                get_belt_number(test.tested_for_belt),  # Convert belt to number
                test_date,
                ("M" if s.gender.lower() == "male" else
                "F" if s.gender.lower() == "female" else
                "O"),
                s.age,
                s.instructor_name or "",
                "",
                "",
                test.examiner_name or "",  # Name of Examiner
                "",
                "",
                "",
                "", #MASTER VENUGOPALAN
                ""   # EMPTY Remarks
            ]

            for col, val in enumerate(data, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = thin
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(bold=True)

            ws.row_dimensions[row].height = 22

        workbook.save(response)
        return response


from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = (JSONParser, MultiPartParser, FormParser)

    @action(detail=False, methods=['get', 'patch'])
    def me(self, request):
        if request.method == 'GET':
            serializer = self.get_serializer(request.user)
            return Response(serializer.data)
        elif request.method == 'PATCH':
            serializer = self.get_serializer(request.user, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def change_password(self, request):
        serializer = PasswordChangeSerializer(data=request.data)
        if serializer.is_valid():
            user = request.user
            new_pass = serializer.data.get('new_password')
            user.set_password(new_pass)
            user.save()
            
            # Save plain password to profile (Unsafe, requested by user)
            from .models import UserProfile
            profile, created = UserProfile.objects.get_or_create(user=user)
            profile.plain_password = new_pass
            profile.save()

            return Response({'status': 'password set'}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class StudentEventViewSet(viewsets.ModelViewSet):
    queryset = StudentEvent.objects.all().order_by('-date')
    serializer_class = StudentEventSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['student']

from rest_framework import serializers
from django.contrib.auth import authenticate
from django.utils.translation import gettext_lazy as _

class EmailAuthTokenSerializer(serializers.Serializer):
    email = serializers.EmailField(label=_("Email"), write_only=True)
    password = serializers.CharField(
        label=_("Password"),
        style={'input_type': 'password'},
        trim_whitespace=False,
        write_only=True
    )
    token = serializers.CharField(
        label=_("Token"),
        read_only=True
    )

    def validate(self, attrs):
        email = attrs.get('email')
        password = attrs.get('password')

        if email and password:
            user = authenticate(request=self.context.get('request'),
                                username=email, password=password)

            # The authenticate call simply returns None for is_active=False
            # users. (Assuming the default ModelBackend authentication
            # backend.)
            if not user:
                msg = _('Unable to log in with provided credentials.')
                raise serializers.ValidationError(msg, code='authorization')
        else:
            msg = _('Must include "email" and "password".')
            raise serializers.ValidationError(msg, code='authorization')

        attrs['user'] = user
        return attrs

class CustomAuthToken(ObtainAuthToken):
    serializer_class = EmailAuthTokenSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data,
                                           context={'request': request})
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        token, created = Token.objects.get_or_create(user=user)
        
        user_role_str = "viewer"
        if user.is_superuser:
            user_role_str = "admin"
        elif hasattr(user, 'user_role'):
            user_role_str = user.user_role.role

        return Response({
            'token': token.key,
            'user_id': user.pk,
            'email': user.email,
            'role': user_role_str
        })

from rest_framework.decorators import api_view, permission_classes
from django.db import IntegrityError

@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def register_user(request):
    email = request.data.get('email')
    password = request.data.get('password')
    
    if not email or not password:
        return Response({'error': 'Email and password are required'}, status=status.HTTP_400_BAD_REQUEST)
        
    try:
        if User.objects.filter(email=email).exists():
             return Response({'error': 'User with this email already exists'}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.create_user(username=email, email=email, password=password)
        # Default role
        UserRole.objects.create(user=user, role='viewer')
        
        token, created = Token.objects.get_or_create(user=user)
        
        return Response({
            'token': token.key,
            'user_id': user.pk,
            'email': user.email,
            'role': 'viewer'
        }, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

